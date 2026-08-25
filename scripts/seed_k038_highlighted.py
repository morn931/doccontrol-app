"""
Seed the K038 documents Jarrod selected for migration into the carry-over register.

⚠️ HIS SELECTION IS THE VISIBLE ROWS, NOT THE HIGHLIGHT. He hides the rows that are not
coming across, and the yellow fill stays on the hidden ones — reading the fill reports 238
where the answer is 212, and would carry documents he has already removed. Row visibility
is the signal. This is why the extractor loads the workbook TWICE: once for styles and row
visibility, once with data_only for the cached values, because the document number is a
formula and reads back as "=A142&B142&..." otherwise.

These rows differ from the transfer-folder ones in two ways:
  · their metadata comes from the K038 CDDL itself, which is authoritative — not from an
    AI reading a title block. It is still loaded into the ai_* (suggestion) columns rather
    than the decision columns, because a controller must still choose it, but the register
    labels it "K038 CDDL" instead of "reader" so nobody mistakes its provenance.
  · their FILES live on SharePoint (ENG2), reached through mddr_entries.file_link, not in
    the OneDrive transfer folder.

    python scripts/seed_k038_highlighted.py <workbook.xlsx>            # dry run
    python scripts/seed_k038_highlighted.py <workbook.xlsx> --apply
"""
import json
import os
import re
import sys
import urllib.error
import urllib.request
import warnings

warnings.filterwarnings("ignore")
from openpyxl import load_workbook

HERE = os.path.dirname(os.path.abspath(__file__))
ENV = {}
for line in open(os.path.join(HERE, "..", ".env.local"), encoding="utf-8", errors="replace"):
    line = line.strip()
    if not line or line.startswith("#") or "=" not in line:
        continue
    k, v = line.split("=", 1)
    v = v.strip().strip('"').strip("'").replace("\\n", "").strip()
    if v:
        ENV[k.strip()] = v
URL = os.environ.get("NEXT_PUBLIC_SUPABASE_URL") or ENV.get("NEXT_PUBLIC_SUPABASE_URL")
KEY = os.environ.get("SUPABASE_SERVICE_ROLE_KEY") or ENV.get("SUPABASE_SERVICE_ROLE_KEY")
if not (URL and KEY):
    sys.exit("NEXT_PUBLIC_SUPABASE_URL / SUPABASE_SERVICE_ROLE_KEY not found")
H = {"apikey": KEY, "Authorization": f"Bearer {KEY}", "Content-Type": "application/json"}
norm = lambda s: re.sub(r"[^A-Z0-9]", "", str(s or "").upper())
SOURCE = "k038 highlighted"

DISCIPLINE = {
    "E": "Electrical", "C": "Civil", "I": "Instrumentation", "M": "Mechanical",
    "F": "Process & Control", "W": "Civil — earthworks", "S": "Structural", "T": "Telecoms",
    "G": "General", "A": "Architectural", "H": "HSE", "Q": "Quality", "P": "Piping",
    "B": "Project Controls", "J": "Environmental", "U": "Utilities",
}


def rest(path, method="GET", body=None, extra=None):
    req = urllib.request.Request(f"{URL}/rest/v1/{path}", method=method,
                                 data=json.dumps(body).encode() if body is not None else None,
                                 headers={**H, **(extra or {})})
    try:
        raw = urllib.request.urlopen(req, timeout=180).read()
    except urllib.error.HTTPError as e:
        # PostgREST puts the real reason in the body; without this a missing column is just
        # "HTTP Error 400: Bad Request" and you go hunting for the wrong thing.
        detail = e.read().decode("utf8", "ignore")[:400]
        raise SystemExit(f"\n  {method} {path.split('?')[0]} failed: {e.code}\n  {detail}\n")
    return json.loads(raw) if raw else []


def page(path):
    out, off = [], 0
    while True:
        r = rest(f"{path}&offset={off}&limit=1000")
        if not r:
            break
        out += r
        if len(r) < 1000:
            break
        off += 1000
    return out


def visible_rows(path):
    """Only the rows Jarrod left visible — see the module docstring for why."""
    wb = load_workbook(path)                    # styles + row visibility
    wbv = load_workbook(path, data_only=True)   # cached values
    ws, wsv = wb["CDDL"], wbv["CDDL"]
    hdr = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    col = lambda n: next((i + 1 for i, h in enumerate(hdr) if str(h).strip() == n), None)
    like = lambda f: next((i + 1 for i, h in enumerate(hdr) if f in str(h)), None)
    C = {
        "doc": hdr.index("RDMC Document Number") + 1, "title": col("Doc Description"),
        "full": col("Full Title"), "status": like("Aconex Doc Status"),
        "review": like("Aconex Review Status"), "disc": col("Discipline"),
        "area": like("Area/WBS Code"), "type": col("Document Type"), "rev": col("Revision"),
        "owner": col("Doc Owner"), "cat": col("Document Category"), "seq": col("Sequence No"),
        "main": col("Main Group"), "sub": col("Sub Group"),
    }
    out = {}
    for r in range(2, ws.max_row + 1):
        if ws.row_dimensions[r].hidden:
            continue
        d = str(wsv.cell(r, C["doc"]).value or "").strip()
        if not d:
            continue
        g = lambda k: str(wsv.cell(r, C[k]).value or "").strip() if C[k] else ""
        out[norm(d)] = {"docno": d, **{k: g(k) for k in C if k != "doc"}}
    return out


def main():
    if len(sys.argv) < 2:
        sys.exit(__doc__)
    path = sys.argv[1]
    apply = "--apply" in sys.argv

    sel = visible_rows(path)
    print(f"visible (Jarrod's selection) : {len(sel)}")

    existing = page("cddl_carryover?select=temp_ref,source_path,legacy_docno,ai_docno,source&order=temp_ref")
    have = {}
    for r in existing:
        for n in (r.get("legacy_docno"), r.get("ai_docno")):
            if n:
                have.setdefault(norm(n), r)
    used = {r["temp_ref"] for r in existing}
    already = [k for k in sel if k in have]
    todo = {k: v for k, v in sel.items() if k not in have}
    print(f"  already in the register    : {len(already)}")
    print(f"  to add                     : {len(todo)}")

    # where the files live
    mddr = page("mddr_entries?select=id,normalized_document_number,document_number,file_link"
                "&package_code=eq.K038&order=id")
    by = {}
    for m in mddr:
        if not m.get("file_link"):
            continue
        for k in (norm(m.get("normalized_document_number")), norm(m.get("document_number"))):
            if k:
                by.setdefault(k, m)
    with_file = sum(1 for k in todo if k in by)
    print(f"  of those, openable         : {with_file}   (no file: {len(todo) - with_file})")

    def next_ref():
        n = 1
        while f"CO-{n:03d}" in used:
            n += 1
        used.add(f"CO-{n:03d}")
        return f"CO-{n:03d}"

    rows = []
    for k, v in sorted(todo.items(), key=lambda kv: kv[1]["docno"]):
        m = by.get(k)
        rows.append({
            "temp_ref": next_ref(),
            "source": SOURCE,
            # source_path is NOT NULL and identifies the document; for these the file lives
            # on SharePoint, so the document number is the stable identifier.
            "source_path": v["docno"],
            "source_files": [],
            "target_package": "K038 CDDL (migrate to K124)",
            "doc_class": v["cat"] or None,
            "legacy_docno": v["docno"],
            "legacy_package": "K038",
            "legacy_area": v["area"] or None,
            "mddr_id": m["id"] if m else None,
            "file_link": m["file_link"] if m else None,
            # Suggestions, from the K038 CDDL rather than from reading the document. They
            # live in the ai_* columns because that is what the register offers beside each
            # field — the label in the UI says where they came from.
            "ai_docno": v["docno"],
            "ai_title": v["title"] or v["full"] or None,
            "ai_revision": v["rev"] or None,
            "ai_status": v["status"] or None,
            "ai_discipline": DISCIPLINE.get(v["disc"].upper(), v["disc"]) or None,
            "ai_doc_type": v["type"] or None,
            "ai_summary": f"From the K038 CDDL: {v['full'] or v['title']}."
                          + (f" Review status {v['review']}." if v["review"] else "")
                          + (f" Owner {v['owner']}." if v["owner"] else ""),
            "ai_confidence": "high",
            "ai_read_at": None if m else None,
            "ai_error": None if m else "No file found for this document in the K038 register",
        })

    if not apply:
        print("\n  sample:")
        for r in rows[:5]:
            print(f"    {r['temp_ref']}  {r['legacy_docno'][:28]:28s} {str(r['ai_status'])[:26]:26s} file={'yes' if r['file_link'] else 'NO'}")
        print(f"\n  (dry run — {len(rows)} rows not written. Re-run with --apply)")
        return

    for i in range(0, len(rows), 100):
        rest("cddl_carryover?on_conflict=temp_ref", "POST", rows[i:i + 100],
             {"Prefer": "resolution=merge-duplicates,return=minimal"})
        print(f"\r  written {min(i + 100, len(rows))}/{len(rows)}", end="")
    total = page("cddl_carryover?select=temp_ref,source&order=temp_ref")
    import collections
    print(f"\n  done — {len(total)} rows in the register")
    for k, v in collections.Counter(r.get("source") or "(none)" for r in total).most_common():
        print(f"     {v:5d}  {k}")


if __name__ == "__main__":
    main()
